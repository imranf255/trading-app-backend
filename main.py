from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import Session, sessionmaker
from datetime import datetime, timedelta
import os
import requests  # ✅ REPLACED yfinance with requests
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse
import io
from pydantic import BaseModel

app = FastAPI()

# 🔑 Finnhub API Key
FINNHUB_API_KEY = "d609vvhr01qgk0vifc80d609vvhr01qgk0vifc8g"

# Market hours configuration (toggleable via API - no restart needed!)
market_hours_config = {"enforce": True}

class TradeRequest(BaseModel):
    ticker: str
    shares: int

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://custommadecake.com.au",
        "https://trading.custommadecake.com.au",
        "http://localhost:3000",
        "http://127.0.0.1:5500"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE_URL = "sqlite:///./trades.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    password = Column(String)
    cash = Column(Float, default=10000.0)

class Trade(Base):
    __tablename__ = "trades"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    ticker = Column(String)
    action = Column(String)
    shares = Column(Integer)
    price = Column(Float)
    total = Column(Float)
    timestamp = Column(DateTime, default=datetime.now)

class PriceAlert(Base):
    __tablename__ = "price_alerts"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    ticker = Column(String)
    original_price = Column(Float)
    current_price = Column(Float)
    change_percent = Column(Float)
    triggered_at = Column(DateTime, default=datetime.now)
    read = Column(Integer, default=0)

class LimitOrder(Base):
    __tablename__ = "limit_orders"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    ticker = Column(String)
    action = Column(String)
    shares = Column(Integer)
    limit_price = Column(Float)
    status = Column(String, default="pending")
    created_at = Column(DateTime, default=datetime.now)
    executed_at = Column(DateTime, nullable=True)

Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def is_market_open():
    ny_tz = pytz.timezone("America/New_York")
    now = datetime.now(ny_tz)
    if now.weekday() >= 5:
        return False
    market_open = now.replace(hour=9, minute=30, second=0, microsecond=0)
    market_close = now.replace(hour=16, minute=0, second=0, microsecond=0)
    return market_open <= now <= market_close

# ✅ NEW: Finnhub price fetcher
def get_stock_price_finnhub(ticker: str) -> float:
    """Fetch real-time stock price from Finnhub API"""
    try:
        url = f"https://finnhub.io/api/v1/quote?symbol={ticker.upper()}&token={FINNHUB_API_KEY}"
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        # Finnhub returns 'c' for current price
        current_price = data.get('c', 0)
        
        if current_price == 0:
            return None
        
        return round(current_price, 2)
    except:
        return None

# ✅ UPDATED: Fallback price using Finnhub
def get_fallback_price(ticker):
    return get_stock_price_finnhub(ticker)

def check_price_alert(db: Session, user_id: int, ticker: str, current_price: float):
    recent_trades = db.query(Trade).filter(Trade.user_id == user_id, Trade.ticker == ticker).order_by(Trade.timestamp.desc()).limit(1).all()
    if recent_trades:
        original_price = recent_trades[0].price
        change_percent = ((current_price - original_price) / original_price) * 100
        if abs(change_percent) > 5:
            alert = PriceAlert(user_id=user_id, ticker=ticker, original_price=original_price, current_price=current_price, change_percent=change_percent)
            db.add(alert)
            db.commit()

# ============ ENHANCED EXCEL DOWNLOAD ============
@app.get("/download-history/{user_id}")
def download_history(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    trades = db.query(Trade).filter(Trade.user_id == user_id).order_by(Trade.timestamp.desc()).all()
    if not trades:
        raise HTTPException(status_code=404, detail="No trades found")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: SUMMARY ==========
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    ws_summary['A1'] = "TRADING ACCOUNT SUMMARY"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = "Username"
    ws_summary['B3'] = user.username
    ws_summary['A4'] = "Report Date"
    ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary['A6'] = "ACCOUNT SNAPSHOT"
    ws_summary['A6'].font = Font(bold=True, size=11)
    ws_summary['A7'] = "Starting Capital"
    ws_summary['B7'] = 10000.00
    ws_summary['A8'] = "Current Cash"
    ws_summary['B8'] = f"{user.cash:.2f}"
    
    total_invested = sum(t.total for t in trades if t.action == "buy")
    total_returned = sum(t.total for t in trades if t.action == "sell")
    net_pl = total_returned - total_invested
    
    ws_summary['A9'] = "Total Invested"
    ws_summary['B9'] = f"{total_invested:.2f}"
    ws_summary['A10'] = "Total Returned"
    ws_summary['B10'] = f"{total_returned:.2f}"
    ws_summary['A11'] = "Net Profit/Loss"
    ws_summary['B11'] = f"{net_pl:.2f}"
    ws_summary['B11'].font = Font(bold=True, color="00B050" if net_pl >= 0 else "FF0000")
    
    ws_summary['A13'] = "TRADING STATISTICS"
    ws_summary['A13'].font = Font(bold=True, size=11)
    
    total_trades = len(trades)
    winning_trades = 0
    losing_trades = 0
    biggest_gain = 0
    biggest_loss = 0
    ticker_trades = {}
    
    for trade in trades:
        if trade.ticker not in ticker_trades:
            ticker_trades[trade.ticker] = []
        ticker_trades[trade.ticker].append(trade)
    
    for ticker, tlist in ticker_trades.items():
        buy_total = sum(t.total for t in tlist if t.action == "buy")
        sell_total = sum(t.total for t in tlist if t.action == "sell")
        if sell_total > 0:
            profit = sell_total - buy_total
            if profit > 0:
                winning_trades += 1
                biggest_gain = max(biggest_gain, profit)
            else:
                losing_trades += 1
                biggest_loss = min(biggest_loss, profit)
    
    win_rate = (winning_trades / (winning_trades + losing_trades) * 100) if (winning_trades + losing_trades) > 0 else 0
    
    ws_summary['A14'] = "Total Trades"
    ws_summary['B14'] = total_trades
    ws_summary['A15'] = "Winning Trades"
    ws_summary['B15'] = winning_trades
    ws_summary['A16'] = "Losing Trades"
    ws_summary['B16'] = losing_trades
    ws_summary['A17'] = "Win Rate"
    ws_summary['B17'] = f"{win_rate:.2f}%"
    ws_summary['A18'] = "Biggest Gain"
    ws_summary['B18'] = f"{biggest_gain:.2f}"
    ws_summary['B18'].font = Font(color="00B050", bold=True)
    ws_summary['A19'] = "Biggest Loss"
    ws_summary['B19'] = f"{biggest_loss:.2f}"
    ws_summary['B19'].font = Font(color="FF0000", bold=True)
    
    # ========== SHEET 2: TRADE HISTORY ==========
    ws_trade = wb.create_sheet("Trade History")
    headers = ["Date", "Ticker", "Action", "Shares", "Price", "Total", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws_trade.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for idx, trade in enumerate(trades, 2):
        ws_trade[f'A{idx}'] = trade.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        ws_trade[f'B{idx}'] = trade.ticker
        ws_trade[f'C{idx}'] = trade.action.upper()
        ws_trade[f'D{idx}'] = trade.shares
        ws_trade[f'E{idx}'] = f"{trade.price:.2f}"
        ws_trade[f'F{idx}'] = f"{trade.total:.2f}"
        ws_trade[f'G{idx}'] = "Completed"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_trade[f'{col}{idx}'].border = border
    
    ws_trade.column_dimensions['A'].width = 20
    ws_trade.column_dimensions['B'].width = 12
    ws_trade.column_dimensions['C'].width = 10
    ws_trade.column_dimensions['D'].width = 10
    ws_trade.column_dimensions['E'].width = 12
    ws_trade.column_dimensions['F'].width = 12
    ws_trade.column_dimensions['G'].width = 12
    
    # ========== SHEET 3: PROFIT & LOSS ==========
    ws_pl = wb.create_sheet("Profit & Loss")
    headers_pl = ["Ticker", "Buy Total", "Sell Total", "Gross P&L", "P&L %", "Status"]
    for col, header in enumerate(headers_pl, 1):
        cell = ws_pl.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    ticker_analysis = {}
    for trade in trades:
        if trade.ticker not in ticker_analysis:
            ticker_analysis[trade.ticker] = {"buy": 0, "sell": 0, "buy_shares": 0, "sell_shares": 0}
        if trade.action == "buy":
            ticker_analysis[trade.ticker]["buy"] += trade.total
            ticker_analysis[trade.ticker]["buy_shares"] += trade.shares
        else:
            ticker_analysis[trade.ticker]["sell"] += trade.total
            ticker_analysis[trade.ticker]["sell_shares"] += trade.shares
    
    total_buy = 0
    total_sell = 0
    row_pl = 2
    
    for ticker, data in ticker_analysis.items():
        buy_total = data["buy"]
        sell_total = data["sell"]
        pl = sell_total - buy_total
        pl_pct = (pl / buy_total * 100) if buy_total > 0 else 0
        status = "✓ Closed" if data["buy_shares"] == data["sell_shares"] else "⊗ Open"
        
        ws_pl[f'A{row_pl}'] = ticker
        ws_pl[f'B{row_pl}'] = f"{buy_total:.2f}"
        ws_pl[f'C{row_pl}'] = f"{sell_total:.2f}"
        ws_pl[f'D{row_pl}'] = f"{pl:.2f}"
        ws_pl[f'D{row_pl}'].font = Font(color="00B050" if pl >= 0 else "FF0000", bold=True)
        ws_pl[f'E{row_pl}'] = f"{pl_pct:.2f}%"
        ws_pl[f'E{row_pl}'].font = Font(color="00B050" if pl_pct >= 0 else "FF0000", bold=True)
        ws_pl[f'F{row_pl}'] = status
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_pl[f'{col}{row_pl}'].border = border
        
        total_buy += buy_total
        total_sell += sell_total
        row_pl += 1
    
    # Total row
    total_pl = total_sell - total_buy
    total_pl_pct = (total_pl / total_buy * 100) if total_buy > 0 else 0
    ws_pl[f'A{row_pl}'] = "TOTAL"
    ws_pl[f'A{row_pl}'].font = Font(bold=True)
    ws_pl[f'B{row_pl}'] = f"{total_buy:.2f}"
    ws_pl[f'B{row_pl}'].font = Font(bold=True)
    ws_pl[f'C{row_pl}'] = f"{total_sell:.2f}"
    ws_pl[f'C{row_pl}'].font = Font(bold=True)
    ws_pl[f'D{row_pl}'] = f"{total_pl:.2f}"
    ws_pl[f'D{row_pl}'].font = Font(bold=True, color="00B050" if total_pl >= 0 else "FF0000")
    ws_pl[f'E{row_pl}'] = f"{total_pl_pct:.2f}%"
    ws_pl[f'E{row_pl}'].font = Font(bold=True, color="00B050" if total_pl_pct >= 0 else "FF0000")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_pl[f'{col}{row_pl}'].border = border
        ws_pl[f'{col}{row_pl}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_pl.column_dimensions['A'].width = 12
    ws_pl.column_dimensions['B'].width = 12
    ws_pl.column_dimensions['C'].width = 12
    ws_pl.column_dimensions['D'].width = 12
    ws_pl.column_dimensions['E'].width = 10
    ws_pl.column_dimensions['F'].width = 12
    
    # ========== SHEET 4: PRICE MOVEMENT ==========
    ws_move = wb.create_sheet("Price Movement")
    headers_move = ["Ticker", "Entry Price", "Current Price", "Move %", "Shares", "Entry Value", "Current Value", "Unrealized P&L"]
    for col, header in enumerate(headers_move, 1):
        cell = ws_move.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    current_positions = {}
    for trade in trades:
        if trade.ticker not in current_positions:
            current_positions[trade.ticker] = {"shares": 0, "avg_price": 0, "total_cost": 0}
        if trade.action == "buy":
            current_positions[trade.ticker]["total_cost"] += trade.total
            current_positions[trade.ticker]["shares"] += trade.shares
        else:
            current_positions[trade.ticker]["shares"] -= trade.shares
    
    row_move = 2
    for ticker, pos in current_positions.items():
        if pos["shares"] > 0:
            avg_price = pos["total_cost"] / pos["shares"]
            current_price = get_stock_price_finnhub(ticker)
            if not current_price:
                current_price = avg_price
            
            move_pct = ((current_price - avg_price) / avg_price * 100) if avg_price > 0 else 0
            entry_value = avg_price * pos["shares"]
            current_value = current_price * pos["shares"]
            unrealized_pl = current_value - entry_value
            
            ws_move[f'A{row_move}'] = ticker
            ws_move[f'B{row_move}'] = f"{avg_price:.2f}"
            ws_move[f'C{row_move}'] = f"{current_price:.2f}"
            ws_move[f'D{row_move}'] = f"{move_pct:.2f}%"
            ws_move[f'D{row_move}'].font = Font(color="00B050" if move_pct >= 0 else "FF0000", bold=True)
            ws_move[f'E{row_move}'] = pos["shares"]
            ws_move[f'F{row_move}'] = f"{entry_value:.2f}"
            ws_move[f'G{row_move}'] = f"{current_value:.2f}"
            ws_move[f'H{row_move}'] = f"{unrealized_pl:.2f}"
            ws_move[f'H{row_move}'].font = Font(color="00B050" if unrealized_pl >= 0 else "FF0000", bold=True)
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws_move[f'{col}{row_move}'].border = border
            
            row_move += 1
    
    ws_move.column_dimensions['A'].width = 12
    ws_move.column_dimensions['B'].width = 12
    ws_move.column_dimensions['C'].width = 14
    ws_move.column_dimensions['D'].width = 10
    ws_move.column_dimensions['E'].width = 10
    ws_move.column_dimensions['F'].width = 12
    ws_move.column_dimensions['G'].width = 14
    ws_move.column_dimensions['H'].width = 15

        # ========== SHEET 5: TRANSACTION DETAILS ==========
    ws_trans = wb.create_sheet("Transaction Detail")
    headers_trans = ["#", "Date", "Ticker", "Type", "Shares", "Price", "Amount", "Running Cash"]
    for col, header in enumerate(headers_trans, 1):
        cell = ws_trans.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    running_cash = user.cash + sum(t.total if t.action == "sell" else -t.total for t in trades)
    row_trans = 2
    
    for idx, trade in enumerate(trades, 1):
        ws_trans[f'A{row_trans}'] = idx
        ws_trans[f'B{row_trans}'] = trade.timestamp.strftime("%Y-%m-%d")
        ws_trans[f'C{row_trans}'] = trade.ticker
        ws_trans[f'D{row_trans}'] = trade.action.upper()
        ws_trans[f'D{row_trans}'].font = Font(color="00B050" if trade.action == "buy" else "FF0000", bold=True)
        ws_trans[f'E{row_trans}'] = trade.shares
        ws_trans[f'F{row_trans}'] = f"{trade.price:.2f}"
        ws_trans[f'G{row_trans}'] = f"{trade.total:.2f}"
        ws_trans[f'H{row_trans}'] = f"{running_cash:.2f}"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_trans[f'{col}{row_trans}'].border = border
        
        running_cash -= (trade.total if trade.action == "sell" else -trade.total)
        row_trans += 1
    
    ws_trans.column_dimensions['A'].width = 6
    ws_trans.column_dimensions['B'].width = 12
    ws_trans.column_dimensions['C'].width = 10
    ws_trans.column_dimensions['D'].width = 10
    ws_trans.column_dimensions['E'].width = 10
    ws_trans.column_dimensions['F'].width = 10
    ws_trans.column_dimensions['G'].width = 12
    ws_trans.column_dimensions['H'].width = 14
    
    # ========== SHEET 6: PERFORMANCE STATS ==========
    ws_stats = wb.create_sheet("Performance Stats")
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    ws_stats['A1'] = "PERFORMANCE METRICS"
    ws_stats['A1'].font = title_font
    ws_stats['A1'].fill = title_fill
    ws_stats.merge_cells('A1:B1')
    
    row_stats = 3
    metrics = [
        ("Total Trades", str(total_trades)),
        ("Winning Trades", str(winning_trades)),
        ("Losing Trades", str(losing_trades)),
        ("Win Rate %", f"{win_rate:.2f}%"),
        ("Biggest Gain", f"${biggest_gain:.2f}"),
        ("Biggest Loss", f"${biggest_loss:.2f}"),
        ("Total Invested", f"${total_invested:.2f}"),
        ("Total Returned", f"${total_returned:.2f}"),
        ("Net Profit/Loss", f"${net_pl:.2f}"),
        ("Return on Investment", f"{(net_pl / 10000 * 100):.2f}%"),
        ("Starting Capital", "$10,000.00"),
        ("Current Cash", f"${user.cash:.2f}"),
    ]
    
    for metric, value in metrics:
        ws_stats[f'A{row_stats}'] = metric
        ws_stats[f'A{row_stats}'].font = Font(bold=True)
        ws_stats[f'B{row_stats}'] = value
        if "Gain" in metric:
            ws_stats[f'B{row_stats}'].font = Font(color="00B050", bold=True)
        elif "Loss" in metric:
            ws_stats[f'B{row_stats}'].font = Font(color="FF0000", bold=True)
        elif "Profit" in metric:
            ws_stats[f'B{row_stats}'].font = Font(color="00B050" if net_pl >= 0 else "FF0000", bold=True)
        row_stats += 1
    
    # ========== SHEET 7: PORTFOLIO ANALYSIS ==========
    ws_port = wb.create_sheet("Portfolio Analysis")
    headers_port = ["Ticker", "Shares", "Avg Cost", "Current Price", "Position Value", "Unrealized P&L", "Portfolio Weight %"]
    for col, header in enumerate(headers_port, 1):
        cell = ws_port.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    total_position_value = 0
    positions_data = []
    
    for ticker, pos in current_positions.items():
        if pos["shares"] > 0:
            avg_price = pos["total_cost"] / pos["shares"]
            current_price = get_stock_price_finnhub(ticker)
            if not current_price:
                current_price = avg_price
            
            position_value = current_price * pos["shares"]
            unrealized_pl = position_value - pos["total_cost"]
            total_position_value += position_value
            
            positions_data.append({
                "ticker": ticker,
                "shares": pos["shares"],
                "avg_price": avg_price,
                "current_price": current_price,
                "position_value": position_value,
                "unrealized_pl": unrealized_pl
            })
    
    row_port = 2
    for pos in positions_data:
        weight = (pos["position_value"] / total_position_value * 100) if total_position_value > 0 else 0
        
        ws_port[f'A{row_port}'] = pos["ticker"]
        ws_port[f'B{row_port}'] = pos["shares"]
        ws_port[f'C{row_port}'] = f"{pos['avg_price']:.2f}"
        ws_port[f'D{row_port}'] = f"{pos['current_price']:.2f}"
        ws_port[f'E{row_port}'] = f"{pos['position_value']:.2f}"
        ws_port[f'F{row_port}'] = f"{pos['unrealized_pl']:.2f}"
        ws_port[f'F{row_port}'].font = Font(color="00B050" if pos['unrealized_pl'] >= 0 else "FF0000", bold=True)
        ws_port[f'G{row_port}'] = f"{weight:.2f}%"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_port[f'{col}{row_port}'].border = border
        
        row_port += 1
    
    # Total row
    ws_port[f'A{row_port}'] = "TOTAL"
    ws_port[f'A{row_port}'].font = Font(bold=True)
    ws_port[f'E{row_port}'] = f"{total_position_value:.2f}"
    ws_port[f'E{row_port}'].font = Font(bold=True)
    ws_port[f'G{row_port}'] = "100.00%"
    ws_port[f'G{row_port}'].font = Font(bold=True)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_port[f'{col}{row_port}'].border = border
        ws_port[f'{col}{row_port}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_port.column_dimensions['A'].width = 12
    ws_port.column_dimensions['B'].width = 10
    ws_port.column_dimensions['C'].width = 12
    ws_port.column_dimensions['D'].width = 14
    ws_port.column_dimensions['E'].width = 14
    ws_port.column_dimensions['F'].width = 15
    ws_port.column_dimensions['G'].width = 17

    
    # Additional sheets (Transaction Detail, Performance Stats, Portfolio Analysis) omitted for brevity but follow same pattern
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trading_report_{user.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# ============ AUTH ENDPOINTS ============
@app.post("/register")
def register(username: str, password: str, db: Session = Depends(get_db)):
    existing_user = db.query(User).filter(User.username == username).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    new_user = User(username=username, password=password, cash=10000.0)
    db.add(new_user)
    db.commit()
    return {"success": True, "message": "User registered successfully"}

@app.post("/login")
def login(username: str, password: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username, User.password == password).first()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"user_id": user.id, "username": user.username, "cash": user.cash}

@app.get("/market-status")
def market_status():
    is_open = is_market_open()
    ny_tz = pytz.timezone("America/New_York")
    now = datetime.now(ny_tz)
    if is_open:
        return {"is_open": True, "status": "Market Open", "message": "NYSE is open until 4:00 PM ET"}
    else:
        if now.weekday() >= 5:
            next_open = "Monday 9:30 AM ET"
        else:
            next_open = "Tomorrow 9:30 AM ET" if now.hour >= 16 else "Today 9:30 AM ET"
        return {"is_open": False, "status": "Market Closed", "message": f"Opens {next_open}"}

# ✅ UPDATED: Price endpoint with Finnhub
@app.get("/price/{ticker}")
def get_price(ticker: str):
    current_price = get_stock_price_finnhub(ticker)
    if current_price is None:
        raise HTTPException(status_code=404, detail=f"Cannot fetch price for {ticker}")
    return {"ticker": ticker, "price": float(current_price)}

# ✅ UPDATED: History endpoint (Finnhub doesn't provide intraday history easily, so simplified)
@app.get("/history/{ticker}")
def get_history(ticker: str, period: str = "1d"):
    # Finnhub's free tier doesn't support intraday history well
    # Return current price as placeholder
    current_price = get_stock_price_finnhub(ticker)
    if not current_price:
        raise HTTPException(status_code=400, detail=f"Cannot fetch history for {ticker}")
    
    # Return simple data (you can enhance this with Finnhub's candle endpoint if needed)
    return {"ticker": ticker, "times": ["09:30", "12:00", "16:00"], "prices": [current_price, current_price, current_price]}

@app.get("/news/{ticker}")
def get_news(ticker: str):
    # Simplified news (Finnhub has news API but requires different endpoint)
    return {"ticker": ticker, "news": [{
        "title": f"{ticker} Market Update",
        "publisher": "Financial News",
        "link": f"https://finance.yahoo.com/quote/{ticker}",
        "published": datetime.now().strftime("%b %d, %Y")
    }]}

@app.post("/buy/{user_id}")
def buy_stock(user_id: int, request: TradeRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if request.shares <= 0:
        raise HTTPException(status_code=400, detail="Shares must be greater than 0")
    
    if market_hours_config["enforce"] and not is_market_open():
        ny_time = datetime.now(pytz.timezone("America/New_York"))
        next_open = "Monday 9:30 AM ET" if ny_time.weekday() >= 5 else "Tomorrow 9:30 AM ET"
        raise HTTPException(status_code=400, detail=f"Market is closed. Opens {next_open}")
    
    current_price = get_stock_price_finnhub(request.ticker)
    if current_price is None:
        raise HTTPException(status_code=400, detail=f"Invalid ticker symbol: {request.ticker}")
    
    total_cost = current_price * request.shares
    commission = total_cost * 0.01  # 1% commission
    total_cost += commission  # Add commission to total cost

    if user.cash < total_cost:
        raise HTTPException(status_code=400, detail=f"Insufficient funds. You need ${total_cost - user.cash:.2f} more")
    
    user.cash -= total_cost
    trade = Trade(user_id=user_id, ticker=request.ticker, action="buy", shares=request.shares, price=current_price, total=total_cost, timestamp=datetime.now())
    db.add(trade)
    check_price_alert(db, user_id, request.ticker, current_price)
    db.commit()
    
    return {"success": True, "message": f"Bought {request.shares} shares of {request.ticker}", "price": current_price, "total": total_cost, "remaining_cash": user.cash}

@app.post("/sell/{user_id}")
def sell_stock(user_id: int, request: TradeRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if request.shares <= 0:
        raise HTTPException(status_code=400, detail="Shares must be greater than 0")
    
    if market_hours_config["enforce"] and not is_market_open():
        ny_time = datetime.now(pytz.timezone("America/New_York"))
        next_open = "Monday 9:30 AM ET" if ny_time.weekday() >= 5 else "Tomorrow 9:30 AM ET"
        raise HTTPException(status_code=400, detail=f"Market is closed. Opens {next_open}")
    
    trades = db.query(Trade).filter(Trade.user_id == user_id, Trade.ticker == request.ticker).all()
    owned_shares = sum(t.shares if t.action == "buy" else -t.shares for t in trades)
    
    if owned_shares < request.shares:
        raise HTTPException(status_code=400, detail=f"You only own {owned_shares} shares of {request.ticker}")
    
    current_price = get_stock_price_finnhub(request.ticker)
    if current_price is None:
        raise HTTPException(status_code=400, detail=f"Invalid ticker symbol: {request.ticker}")
    
    total_value = current_price * request.shares
    commission = total_value * 0.01  # 1% commission
    total_value -= commission  # Deduct commission from proceeds
    user.cash += total_value

    trade = Trade(user_id=user_id, ticker=request.ticker, action="sell", shares=request.shares, price=current_price, total=total_value, timestamp=datetime.now())
    db.add(trade)
    check_price_alert(db, user_id, request.ticker, current_price)
    db.commit()
    
    return {"success": True, "message": f"Sold {request.shares} shares of {request.ticker}", "price": current_price, "total": total_value, "new_cash": user.cash}

@app.get("/portfolio/{user_id}")
def get_portfolio(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    trades = db.query(Trade).filter(Trade.user_id == user_id).all()
    positions = {}
    
    for trade in trades:
        if trade.ticker not in positions:
            positions[trade.ticker] = {"shares": 0, "value": 0}
        positions[trade.ticker]["shares"] += trade.shares if trade.action == "buy" else -trade.shares
    
    positions = {k: v for k, v in positions.items() if v["shares"] > 0}
    
    for ticker in positions:
        current_price = get_fallback_price(ticker)
        if current_price:
            positions[ticker]["price"] = current_price
            positions[ticker]["value"] = current_price * positions[ticker]["shares"]
        else:
            positions[ticker]["price"] = 0
            positions[ticker]["value"] = 0
    
    total_value = sum(p["value"] for p in positions.values())
    return {"cash": user.cash, "positions": positions, "total_account_value": user.cash + total_value, "profit": user.cash + total_value - 10000}

@app.get("/trade-history/{user_id}")
def get_trade_history(user_id: int, db: Session = Depends(get_db)):
    trades = db.query(Trade).filter(Trade.user_id == user_id).order_by(Trade.timestamp.desc()).all()
    return {"trades": [{"id": t.id, "ticker": t.ticker, "action": t.action, "shares": t.shares, "price": t.price, "total": t.total, "timestamp": t.timestamp.isoformat()} for t in trades]}

@app.get("/user-stats/{user_id}")
def get_user_stats(user_id: int, db: Session = Depends(get_db)):
    trades = db.query(Trade).filter(Trade.user_id == user_id).all()
    if not trades:
        return {"total_trades": 0, "win_rate": 0, "biggest_gain": 0, "biggest_loss": 0, "most_traded_stock": None, "total_volume": 0}
    
    ticker_trades = {}
    for trade in trades:
        if trade.ticker not in ticker_trades:
            ticker_trades[trade.ticker] = []
        ticker_trades[trade.ticker].append(trade)
    
    profitable_trades = 0
    biggest_gain = 0
    biggest_loss = 0
    
    for ticker, ticker_list in ticker_trades.items():
        buy_total = sum(t.total for t in ticker_list if t.action == "buy")
        sell_total = sum(t.total for t in ticker_list if t.action == "sell")
        if sell_total > 0:
            profit = sell_total - buy_total
            if profit > 0:
                profitable_trades += 1
                biggest_gain = max(biggest_gain, profit)
            else:
                biggest_loss = min(biggest_loss, profit)
    
    return {"total_trades": len(trades), "win_rate": round(profitable_trades / len(ticker_trades) * 100 if ticker_trades else 0, 2), "biggest_gain": round(biggest_gain, 2), "biggest_loss": round(abs(biggest_loss), 2), "most_traded_stock": max(ticker_trades, key=lambda x: len(ticker_trades[x])) if ticker_trades else None, "total_volume": round(sum(t.total for t in trades), 2)}

@app.get("/leaderboard")
def get_leaderboard(db: Session = Depends(get_db)):
    users = db.query(User).all()
    rankings = []
    
    for user in users:
        trades = db.query(Trade).filter(Trade.user_id == user.id).all()
        positions = {}
        for trade in trades:
            if trade.ticker not in positions:
                positions[trade.ticker] = 0
            positions[trade.ticker] += trade.shares if trade.action == "buy" else -trade.shares
        
        positions = {k: v for k, v in positions.items() if v > 0}
        portfolio_value = 0
        
        for ticker, shares in positions.items():
            price = get_fallback_price(ticker)
            if price:
                portfolio_value += price * shares
        
        rankings.append({"username": user.username, "total_value": user.cash + portfolio_value, "profit": user.cash + portfolio_value - 10000})
    
    rankings.sort(key=lambda x: x["total_value"], reverse=True)
    return {"rankings": rankings}

@app.post("/limit-order/{user_id}")
def create_limit_order(user_id: int, ticker: str, action: str, shares: int, limit_price: float, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if action not in ["buy", "sell"]:
        raise HTTPException(status_code=400, detail="Action must be buy or sell")
    
    limit_order = LimitOrder(user_id=user_id, ticker=ticker, action=action, shares=shares, limit_price=limit_price, status="pending")
    db.add(limit_order)
    db.commit()
    
    return {"success": True, "order_id": limit_order.id, "message": f"Limit order created: {action.upper()} {shares} {ticker} @ {limit_price}"}

@app.get("/limit-orders/{user_id}")
def get_limit_orders(user_id: int, db: Session = Depends(get_db)):
    orders = db.query(LimitOrder).filter(LimitOrder.user_id == user_id, LimitOrder.status == "pending").all()
    return {"orders": [{"id": o.id, "ticker": o.ticker, "action": o.action, "shares": o.shares, "limit_price": o.limit_price, "created_at": o.created_at.isoformat()} for o in orders]}

@app.post("/limit-orders/{order_id}/cancel")
def cancel_limit_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(LimitOrder).filter(LimitOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.status != "pending":
        raise HTTPException(status_code=400, detail="Order already executed or cancelled")
    
    order.status = "cancelled"
    db.commit()
    return {"success": True, "message": "Limit order cancelled"}

@app.get("/portfolio-analytics/{user_id}")
def get_portfolio_analytics(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    trades = db.query(Trade).filter(Trade.user_id == user_id).all()
    positions = {}
    
    for trade in trades:
        if trade.ticker not in positions:
            positions[trade.ticker] = {"shares": 0, "cost_basis": 0}
        if trade.action == "buy":
            positions[trade.ticker]["shares"] += trade.shares
            positions[trade.ticker]["cost_basis"] += trade.total
        else:
            positions[trade.ticker]["shares"] -= trade.shares
            positions[trade.ticker]["cost_basis"] -= trade.total
    
    positions = {k: v for k, v in positions.items() if v["shares"] > 0}
    portfolio_data = []
    total_value = 0
    total_cost = 0
    
    for ticker, pos in positions.items():
        current_price = get_fallback_price(ticker)
        if current_price:
            current_value = current_price * pos["shares"]
            portfolio_data.append({"ticker": ticker, "shares": pos["shares"], "current_value": round(current_value, 2), "profit_loss": round(current_value - pos["cost_basis"], 2)})
            total_value += current_value
            total_cost += pos["cost_basis"]
    
    return {"account_value": round(user.cash + total_value, 2), "cash": round(user.cash, 2), "invested": round(total_value, 2), "total_profit_loss": round(total_value - total_cost, 2), "positions": portfolio_data, "position_count": len(portfolio_data)}

@app.get("/market-hours-config")
def get_market_hours_config():
    return {"enforce_market_hours": market_hours_config["enforce"]}

@app.post("/market-hours-config")
def set_market_hours_config(enforce: bool):
    market_hours_config["enforce"] = enforce
    status = "ENABLED" if enforce else "DISABLED"
    return {"success": True, "enforce_market_hours": enforce, "message": f"Market hours enforcement status: {status}"}

@app.post("/check-limit-orders")
def check_limit_orders(db: Session = Depends(get_db)):
    pending_orders = db.query(LimitOrder).filter(LimitOrder.status == "pending").all()
    executed = 0
    
    for order in pending_orders:
        current_price = get_fallback_price(order.ticker)
        if current_price:
            should_execute = False
            if order.action == "buy" and current_price <= order.limit_price:
                should_execute = True
            elif order.action == "sell" and current_price >= order.limit_price:
                should_execute = True
            
            if should_execute:
                user = db.query(User).filter(User.id == order.user_id).first()
                if order.action == "buy" and user.cash >= (current_price * order.shares):
                    user.cash -= current_price * order.shares
                    trade = Trade(user_id=order.user_id, ticker=order.ticker, action="buy", shares=order.shares, price=current_price, total=current_price * order.shares, timestamp=datetime.now())
                    db.add(trade)
                    order.status = "executed"
                    order.executed_at = datetime.now()
                    executed += 1
                elif order.action == "sell":
                    trades = db.query(Trade).filter(Trade.user_id == order.user_id, Trade.ticker == order.ticker).all()
                    owned_shares = sum(t.shares if t.action == "buy" else -t.shares for t in trades)
                    if owned_shares >= order.shares:
                        user.cash += current_price * order.shares
                        trade = Trade(user_id=order.user_id, ticker=order.ticker, action="sell", shares=order.shares, price=current_price, total=current_price * order.shares, timestamp=datetime.now())
                        db.add(trade)
                        order.status = "executed"
                        order.executed_at = datetime.now()
                        executed += 1
    
    db.commit()
    return {"success": True, "executed": executed}

@app.get("/alerts/{user_id}")
def get_alerts(user_id: int, db: Session = Depends(get_db)):
    alerts = db.query(PriceAlert).filter(PriceAlert.user_id == user_id, PriceAlert.read == 0).order_by(PriceAlert.triggered_at.desc()).all()
    return {"alerts": [{"id": a.id, "ticker": a.ticker, "change_percent": a.change_percent} for a in alerts]}

@app.get("/day-trades/{user_id}")
def get_day_trades(user_id: int, db: Session = Depends(get_db)):
    return {"day_trades_used": 0, "day_trades_remaining": 3, "limit_reached": False}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 10000))
    uvicorn.run(app, host="0.0.0.0", port=port)
